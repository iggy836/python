import openpyxl
from openpyxl.styles import Font, PatternFill

def create_sales_excel():
    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "月間売上"  # シート名を設定

    # ヘッダーを設定
    headers = ['商品名', '単価', '販売数', '合計']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        # ヘッダーを太字に設定
        cell.font = Font(bold=True)
        # 背景色を薄い灰色に設定
        cell.fill = PatternFill(fgColor='EEEEEE', fill_type='solid')

    # 商品データを設定
    products = [
        ['コーヒー', 300, 150],
        ['紅茶', 250, 120],
        ['ジュース', 200, 200]
    ]

    # データを入力
    for row, product in enumerate(products, 2):
        # 商品名、単価、販売数を入力
        for col, value in enumerate(product, 1):
            sheet.cell(row=row, column=col, value=value)
        
        # 合計の計算式を入力
        sheet.cell(row=row, column=4, value=f'=B{row}*C{row}')

    # 列幅を自動調整
    for col in ['A', 'B', 'C', 'D']:
        sheet.column_dimensions[col].width = 15

    # ファイルを保存
    wb.save('売上データ.xlsx')
    print("売上データ.xlsxが作成されました。")

if __name__ == "__main__":
    create_sales_excel()